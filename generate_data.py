import io
import os
import re

import openpyxl
import pandas as pd
import requests
import yaml

EXCEL_URL = "https://www.pbo.gov.au/sites/default/files/2025-04/PBO%20Historical%20fiscal%20data%20-%202025-26%20Budget%20update.xlsx"
BUDGET_SHEET_NAME = "Table 7"
REVENUE_SHEET_NAME = "Table 6"
START_DATA_ROW = 4  # 0-based index, row 5 in Excel


def should_exclude(name: str, unit: str = None) -> bool:
    for excluded_phrases in [
        "Total",
        "total",
        "(underlying basis)",
    ]:
        if excluded_phrases in name:
            return True

    if unit is not None and unit.strip().upper() == "ASL":
        return True
    return False


def fetch_excel_bytes(url: str) -> io.BytesIO:
    """
    Download the Excel file from the given URL and return its bytes as a BytesIO object.
    """
    response = requests.get(url)
    response.raise_for_status()
    return io.BytesIO(response.content)


def build_budget_tree_from_excel(
    excel_bytes: io.BytesIO, years: list[str], sheet_name: str
) -> tuple[dict, list[str]]:
    """
    Parse the Excel sheet and build a hierarchical tree using indent levels.
    Loads the Excel workbook, finds the header row, identifies year columns, and
    builds a nested dictionary structure representing the budget hierarchy.
    """

    # Load Excel workbook and worksheet
    wb = openpyxl.load_workbook(excel_bytes, data_only=True)
    ws = wb[sheet_name]

    # Find the header row containing year labels
    year_pattern = re.compile(r"^\d{4}-\d{2,4}$")
    header_row_idx = None
    year_col_indices = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
        if len([1 for val in row if val and year_pattern.match(str(val).strip())]) > 3:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise RuntimeError(
            "Could not find header row with year labels in first 20 rows"
        )
    headers = [cell.value for cell in ws[header_row_idx + 1]]

    # Identify which columns correspond to which years
    for idx, val in enumerate(headers):
        sval = str(val).strip() if val else ""
        if sval in years:
            year_col_indices.append(idx)
    # Build the hierarchical tree from the Excel rows
    data_start = header_row_idx + 1
    root = {"name": "Australian Government Budget", "children": []}
    stack = []
    for row in ws.iter_rows(min_row=data_start, values_only=False):
        cell = row[0]
        if cell.value is None or str(cell.value).strip().lower().startswith("nan"):
            continue
        name = str(cell.value).strip()
        unit_cell = row[1]
        unit = str(unit_cell.value).strip() if unit_cell and unit_cell.value else "$m"
        if should_exclude(name, unit):
            continue
        indent = (
            cell.alignment.indent if cell.alignment and cell.alignment.indent else 0
        )
        unit_multipliers = {
            "$t": 1_000_000_000_000,
            "$m": 1_000_000,
            "$b": 1_000_000_000,
            "$k": 1_000,
        }
        multiplier = unit_multipliers.get(unit, 1)
        budgets = {}
        for idx, y in zip(year_col_indices, years):
            val = row[idx].value if idx < len(row) else None
            if val is not None:
                try:
                    budgets[y] = float(val) * multiplier
                except Exception:
                    pass
        node = {"name": name, "unit": unit}
        if budgets:
            node["budget"] = budgets
        node["children"] = []
        while stack and stack[-1][0] >= indent:
            stack.pop()
        if not stack:
            root["children"].append(node)
        else:
            stack[-1][1]["children"].append(node)
        stack.append((indent, node))
    return root, years


def export_yearly_yaml(
    tree: dict, years: list[str], output_dir: str, prefix="year_"
) -> None:
    """
    Export each year's budget data to a separate YAML file, filtering out unwanted nodes.
    Recursively filters the tree for each year and writes the result to a YAML file in the output directory.
    """
    os.makedirs(output_dir, exist_ok=True)

    def filter_for_year(node: dict, year: str, unit: str = None) -> dict | None:
        """
        Recursively filter the tree for a specific year, removing excluded nodes and keeping only relevant budget data.
        If a node's value is simply the sum of its children's values, omit the value from the parent node.
        """
        if should_exclude(node["name"], node.get("unit", unit)):
            return None
        new_node = {"name": node["name"]}
        children = []
        if node.get("children"):
            children = [
                filter_for_year(child, year, child.get("unit", unit))
                for child in node["children"]
            ]
            children = [
                c for c in children if c and ("budget" in c or c.get("children"))
            ]
            if children:
                new_node["children"] = children
        # Only add budget if not just the sum of children
        if "budget" in node and year in node["budget"]:
            parent_val = node["budget"][year]
            # Check if children all have budget values for this year
            if children and all("budget" in c and isinstance(c["budget"], (int, float)) for c in children):
                children_sum = sum(c["budget"] for c in children if "budget" in c and isinstance(c["budget"], (int, float)))
                # If parent's value is not exactly the sum of its children, keep it
                if abs(parent_val - children_sum) > 1e-6:
                    new_node["budget"] = parent_val
            else:
                new_node["budget"] = parent_val
        return new_node

    def add_siblings(node: dict) -> None:
        """
        Add a _siblings key to each child node containing a list of its siblings.
        This is used for possible future processing or filtering that may require sibling context.
        """
        children = node.get("children", [])
        for i, child in enumerate(children):
            child["_siblings"] = [c for j, c in enumerate(children) if j != i]
            add_siblings(child)

    add_siblings(tree)

    # Export YAML for each year
    for year in years:
        out = filter_for_year(tree, year)
        out["name"] = f"Australian Government Budget {year}"
        yml_path = os.path.join(output_dir, f"{prefix}{year}.yml")
        with open(yml_path, "w") as f:
            yaml.dump(out, f, sort_keys=False, allow_unicode=True)


if __name__ == "__main__":
    """
    Main entry point for the script. Downloads the Excel file, parses it, builds the budget tree,
    and exports yearly YAML files to the _data directory.
    """
    print("[INFO] Starting data generation...")

    # Download Excel file
    excel_bytes = fetch_excel_bytes(EXCEL_URL)

    # --- Budget ---
    # Read the Excel file with pandas to extract year labels for budget
    raw_budget = pd.read_excel(excel_bytes, sheet_name=BUDGET_SHEET_NAME, header=None)
    year_label_row_budget = list(raw_budget.iloc[2])
    years_budget = []
    for val in year_label_row_budget:
        sval = str(val).strip()
        if len(sval) == 7 and sval[:4].isdigit() and sval[4] == "-":
            years_budget.append(sval)
        elif len(sval) == 4 and sval.isdigit():
            years_budget.append(sval)
    years_budget = sorted(set(years_budget))

    # Build the budget tree for budget
    tree_budget, years_budget = build_budget_tree_from_excel(
        excel_bytes, years_budget, BUDGET_SHEET_NAME
    )

    # Set output directory for budget
    output_dir_budget = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "_data", "budget")
    )

    # Export YAML files for each year (budget)
    export_yearly_yaml(tree_budget, years_budget, output_dir_budget)

    print(
        f"[INFO] Budget data generation complete. YAML files written to {output_dir_budget}"
    )

    # --- Revenue ---
    # Read the Excel file with pandas to extract year labels for revenue
    raw_rev = pd.read_excel(excel_bytes, sheet_name=REVENUE_SHEET_NAME, header=None)
    year_label_row_rev = list(raw_rev.iloc[2])
    years_rev = []
    for val in year_label_row_rev:
        sval = str(val).strip()
        if len(sval) == 7 and sval[:4].isdigit() and sval[4] == "-":
            years_rev.append(sval)
        elif len(sval) == 4 and sval.isdigit():
            years_rev.append(sval)
    years_rev = sorted(set(years_rev))

    # Build the budget tree for revenue
    tree_rev, years_rev = build_budget_tree_from_excel(
        excel_bytes, years_rev, REVENUE_SHEET_NAME
    )

    # --- Special case: Net income tax (Gross income tax withholding - Individuals refunds) ---
    def combine_net_income_tax(node: dict) -> None:
        if not node or not node.get("children"):
            return
        for child in node["children"]:
            combine_net_income_tax(child)
        # Look for the 'Individuals and other withholding taxes' node
        for child in node["children"]:
            if child["name"] == "Individuals and other withholding taxes" and child.get(
                "children"
            ):
                gross = None
                refunds = None
                gross_idx = refunds_idx = None
                for idx, sub in enumerate(child["children"]):
                    if sub["name"] == "Gross income tax withholding":
                        gross = sub
                        gross_idx = idx
                    elif sub["name"] == "Individuals refunds":
                        refunds = sub
                        refunds_idx = idx
                if gross and refunds:
                    # For each year, subtract refunds from gross
                    net_budgets = {}
                    for y in years_rev:
                        gval = gross.get("budget", {}).get(y, 0)
                        rval = refunds.get("budget", {}).get(y, 0)
                        net_budgets[y] = gval + rval  # refunds is negative
                    net_node = {"name": "Net income tax", "budget": net_budgets}
                    # Remove both and insert net
                    new_children = [
                        sub
                        for i, sub in enumerate(child["children"])
                        if i not in (gross_idx, refunds_idx)
                    ]
                    new_children.append(net_node)
                    child["children"] = new_children

    combine_net_income_tax(tree_rev)

    # Set output directory for revenue
    output_dir_rev: str = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "_data", "revenue")
    )

    # Export YAML files for each year (revenue)
    export_yearly_yaml(tree_rev, years_rev, output_dir_rev)

    print(
        f"[INFO] Revenue data generation complete. YAML files written to {output_dir_rev}"
    )
